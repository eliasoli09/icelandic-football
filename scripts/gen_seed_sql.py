#!/usr/bin/env python3
"""Generate SQL seed files from data/seed/ into data/seed/sql/.

Outputs:
  01_teams.sql            all team names seen in history + SofaScore mapping
  02_matches.sql          1,968 historical matches (2019-2025)
  03_sofascore_teams.sql  team stats jsonb (2026 snapshot)
  04_sofascore_players.sql top-150 player snapshot (2026)
"""
import csv, json, pathlib, re

ROOT = pathlib.Path(__file__).resolve().parent.parent
SEED = ROOT / 'data' / 'seed'
OUT = SEED / 'sql'
OUT.mkdir(exist_ok=True)

SOFASCORE_TO_KSI = {
    'Breidablik Kópavogur': 'Breiðablik', 'KR Reykjavík': 'KR',
    'Víkingur Reykjavík': 'Víkingur R.', 'Stjarnan Garðabær': 'Stjarnan',
    'Keflavík IF': 'Keflavík', 'ÍA Akranes': 'ÍA', 'KA Akureyri': 'KA',
    'ÍBV Vestmannaeyjar': 'ÍBV', 'Fram Reykjavík': 'Fram',
    'Þór Akureyri': 'Þór', 'FH Hafnarfjörður': 'FH', 'Valur Reykjavík': 'Valur',
}
# sofascore_team_stats_2026.json uses KSÍ-style names already
q = lambda s: "'" + str(s).replace("'", "''") + "'"

def main():
    rows = list(csv.DictReader(open(SEED / 'matches_hist.csv')))
    teams = sorted({r['home'] for r in rows} | {r['away'] for r in rows})
    tid = {t: i + 1 for i, t in enumerate(teams)}
    ksi_to_sofa = {v: k for k, v in SOFASCORE_TO_KSI.items()}

    with open(OUT / '01_teams.sql', 'w') as f:
        f.write('insert into teams (id, name, sofascore_name) values\n')
        vals = [
            f"({tid[t]}, {q(t)}, {q(ksi_to_sofa[t]) if t in ksi_to_sofa else 'null'})"
            for t in teams
        ]
        f.write(',\n'.join(vals) + '\non conflict (name) do nothing;\n')
        f.write(f"select setval('teams_id_seq', {len(teams) + 1});\n")

    with open(OUT / '02_matches.sql', 'w') as f:
        f.write('insert into matches (id, season, league, phase, home_team, away_team, home_goals, away_goals, status, corrected) values\n')
        vals = [
            f"({r['ksi_id']}, {r['season']}, {q(r['league'])}, {q(r['phase'])}, "
            f"{tid[r['home']]}, {tid[r['away']]}, {r['hg']}, {r['ag']}, 'played', {str(bool(int(r['corrected']))).lower()})"
            for r in rows
        ]
        f.write(',\n'.join(vals) + '\non conflict (id) do nothing;\n')

    stats = json.load(open(SEED / 'sofascore_team_stats_2026.json'))
    with open(OUT / '03_sofascore_teams.sql', 'w') as f:
        vals = []
        for team, s in stats.items():
            if team.startswith('_'):
                continue
            vals.append(f"({tid[team]}, 2026, {q(json.dumps(s, ensure_ascii=False))}::jsonb)")
        f.write('insert into sofascore_team_stats (team_id, season, stats) values\n')
        f.write(',\n'.join(vals) + '\non conflict (team_id, season) do update set stats = excluded.stats, loaded_at = now();\n')

    from openpyxl import load_workbook
    wb = load_workbook(SEED / 'Besta_deild_karla_Top_150_Players.xlsx', read_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(header)}
    with open(OUT / '04_sofascore_players.sql', 'w') as f:
        vals = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx['Player']] is None:
                continue
            extra = {h: row[i] for h, i in idx.items()
                     if h not in ('Rank', 'Player', 'Team', 'Rating', 'Appearances', 'Goals', 'Assists')}
            vals.append(
                f"(2026, {row[idx['Rank']]}, {q(row[idx['Player']])}, {q(SOFASCORE_TO_KSI.get(row[idx['Team']], row[idx['Team']]))}, "
                f"{row[idx['Rating']]}, {row[idx['Appearances']]}, {row[idx['Goals']]}, {row[idx['Assists']]}, "
                f"{q(json.dumps(extra, ensure_ascii=False, default=str))}::jsonb)"
            )
        f.write('insert into sofascore_players (season, rank, name, team, rating, appearances, goals, assists, extra) values\n')
        f.write(',\n'.join(vals) + '\non conflict (season, name) do update set rank = excluded.rank, rating = excluded.rating, appearances = excluded.appearances, goals = excluded.goals, assists = excluded.assists, extra = excluded.extra, loaded_at = now();\n')

    print('teams:', len(teams), '| matches:', len(rows), '| files in', OUT)

if __name__ == '__main__':
    main()
